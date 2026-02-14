"""
PwrSysPro Analysis Suite - Excel Export Module
Exports project data and analysis results to Excel spreadsheets.

Purpose:
- Equipment lists in Excel format
- Cable schedules with calculations
- Analysis results worksheets
- Customizable reports

Dependencies: openpyxl
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not available. Excel export will not work.")
    print("Install with: pip install openpyxl")

from typing import List, Dict, Optional
import io
from datetime import datetime


class ExcelExporter:
    """
    Export project data to Excel workbooks.
    
    Supports multiple export types:
    - Equipment lists
    - Cable schedules
    - Calculation worksheets
    - Summary reports
    """
    
    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export")
        
        self.wb = None
        
        # Styling
        self.header_fill = PatternFill(
            start_color="1F4E78",
            end_color="1F4E78",
            fill_type="solid"
        )
        self.header_font = Font(color="FFFFFF", bold=True, size=11)
        self.title_font = Font(bold=True, size=14)
        
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def export_equipment_list(
        self,
        project_data: Dict,
        component_data: Dict
    ) -> io.BytesIO:
        """
        Export comprehensive equipment list to Excel.
        
        Workbook structure:
        - Summary sheet
        - Breakers sheet
        - Transformers sheet
        - Motors sheet
        - Cables sheet
        
        Args:
            project_data: Project and node data
            component_data: Component library data
        
        Returns:
            BytesIO object containing Excel file
        """
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet
        
        # Extract equipment by type
        nodes = project_data.get('nodes', [])
        connections = project_data.get('connections', [])
        
        breakers = [n for n in nodes if n.get('type') == 'Breaker']
        transformers = [n for n in nodes if n.get('type') == 'Transformer']
        motors = [n for n in nodes if n.get('type') == 'Motor']
        
        # Create sheets
        self._create_equipment_summary_sheet(project_data, nodes)
        
        if breakers:
            self._create_breakers_sheet(breakers, component_data)
        
        if transformers:
            self._create_transformers_sheet(transformers, component_data)
        
        if motors:
            self._create_motors_sheet(motors, component_data)
        
        if connections:
            self._create_cables_sheet(connections, nodes, component_data)
        
        # Save to BytesIO
        output = io.BytesIO()
        self.wb.save(output)
        output.seek(0)
        
        return output
    
    def _create_equipment_summary_sheet(
        self,
        project_data: Dict,
        nodes: List[Dict]
    ):
        """Create summary sheet"""
        ws = self.wb.create_sheet("Summary")
        
        # Project info
        ws['A1'] = f"Equipment List - {project_data.get('name', 'Project')}"
        ws['A1'].font = self.title_font
        
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        
        # Summary table
        row = 4
        ws[f'A{row}'] = "Equipment Type"
        ws[f'B{row}'] = "Count"
        self._apply_header_style(ws, row, 2)
        
        # Count by type
        from collections import Counter
        type_counts = Counter(n.get('type') for n in nodes)
        
        row += 1
        for eq_type, count in sorted(type_counts.items()):
            ws[f'A{row}'] = eq_type
            ws[f'B{row}'] = count
            row += 1
        
        ws[f'A{row}'] = "TOTAL"
        ws[f'B{row}'] = len(nodes)
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        
        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
    
    def _create_breakers_sheet(
        self,
        breakers: List[Dict],
        component_data: Dict
    ):
        """Create breakers sheet"""
        ws = self.wb.create_sheet("Breakers")
        
        # Headers
        headers = [
            'Tag', 'Type', 'Manufacturer', 'Model',
            'Voltage (kV)', 'Rating (A)', 'SC Rating (kA)',
            'Poles', 'Location', 'Status'
        ]
        
        self._write_headers(ws, headers)
        
        # Data
        for i, breaker in enumerate(breakers, start=2):
            props = breaker.get('properties', {})
            comp_id = str(breaker.get('component_library_id', ''))
            comp = component_data.get(comp_id, {})
            
            ws[f'A{i}'] = breaker.get('custom_tag', f"CB-{i}")
            ws[f'B{i}'] = breaker.get('type', 'Breaker')
            ws[f'C{i}'] = comp.get('manufacturer', 'N/A')
            ws[f'D{i}'] = comp.get('model', 'N/A')
            ws[f'E{i}'] = comp.get('voltage_rating', 0)
            ws[f'F{i}'] = comp.get('ampacity_base', 0)
            ws[f'G{i}'] = comp.get('short_circuit_rating', 0)
            ws[f'H{i}'] = props.get('poles', 3)
            ws[f'I{i}'] = props.get('location', 'N/A')
            ws[f'J{i}'] = props.get('status', 'OK')
        
        self._auto_adjust_columns(ws)
    
    def _create_transformers_sheet(
        self,
        transformers: List[Dict],
        component_data: Dict
    ):
        """Create transformers sheet"""
        ws = self.wb.create_sheet("Transformers")
        
        headers = [
            'Tag', 'Manufacturer', 'Model', 'Rating (kVA)',
            'Primary (kV)', 'Secondary (kV)', 'Impedance (%)',
            'Connection', 'Cooling', 'Location'
        ]
        
        self._write_headers(ws, headers)
        
        for i, xfmr in enumerate(transformers, start=2):
            props = xfmr.get('properties', {})
            comp_id = str(xfmr.get('component_library_id', ''))
            comp = component_data.get(comp_id, {})
            
            ws[f'A{i}'] = xfmr.get('custom_tag', f"XFMR-{i}")
            ws[f'B{i}'] = comp.get('manufacturer', 'N/A')
            ws[f'C{i}'] = comp.get('model', 'N/A')
            ws[f'D{i}'] = props.get('rating_kva', 0)
            ws[f'E{i}'] = props.get('primary_voltage', 0)
            ws[f'F{i}'] = props.get('secondary_voltage', 0)
            ws[f'G{i}'] = comp.get('impedance_z_percent', 0)
            ws[f'H{i}'] = props.get('connection', 'Dyn11')
            ws[f'I{i}'] = props.get('cooling', 'ONAN')
            ws[f'J{i}'] = props.get('location', 'N/A')
        
        self._auto_adjust_columns(ws)
    
    def _create_motors_sheet(
        self,
        motors: List[Dict],
        component_data: Dict
    ):
        """Create motors sheet"""
        ws = self.wb.create_sheet("Motors")
        
        headers = [
            'Tag', 'Manufacturer', 'HP/kW', 'Voltage (V)',
            'Current (A)', 'Power Factor', 'Efficiency (%)',
            'Starting Method', 'Application'
        ]
        
        self._write_headers(ws, headers)
        
        for i, motor in enumerate(motors, start=2):
            props = motor.get('properties', {})
            comp_id = str(motor.get('component_library_id', ''))
            comp = component_data.get(comp_id, {})
            
            ws[f'A{i}'] = motor.get('custom_tag', f"MTR-{i}")
            ws[f'B{i}'] = comp.get('manufacturer', 'N/A')
            ws[f'C{i}'] = props.get('hp', 0)
            ws[f'D{i}'] = props.get('voltage', 0)
            ws[f'E{i}'] = props.get('current', 0)
            ws[f'F{i}'] = props.get('power_factor', 0.85)
            ws[f'G{i}'] = props.get('efficiency', 90)
            ws[f'H{i}'] = props.get('starting_method', 'DOL')
            ws[f'I{i}'] = props.get('application', 'General')
        
        self._auto_adjust_columns(ws)
    
    def _create_cables_sheet(
        self,
        connections: List[Dict],
        nodes: List[Dict],
        component_data: Dict
    ):
        """Create cable schedule sheet"""
        ws = self.wb.create_sheet("Cable Schedule")
        
        headers = [
            'Cable Tag', 'From', 'To', 'Type', 'Size (mmÂ²)',
            'Length (m)', 'Cores', 'Voltage (kV)',
            'Installation Method', 'Ampacity (A)',
            'Material', 'Insulation'
        ]
        
        self._write_headers(ws, headers)
        
        # Create node lookup
        node_lookup = {n['id']: n for n in nodes}
        
        for i, conn in enumerate(connections, start=2):
            comp_id = str(conn.get('cable_library_id', ''))
            comp = component_data.get(comp_id, {})
            
            # Get node tags
            from_node = node_lookup.get(conn.get('source_node_id'), {})
            to_node = node_lookup.get(conn.get('target_node_id'), {})
            
            from_tag = from_node.get('custom_tag', 'N/A')
            to_tag = to_node.get('custom_tag', 'N/A')
            
            ws[f'A{i}'] = f"CABLE-{from_tag}-{to_tag}"
            ws[f'B{i}'] = from_tag
            ws[f'C{i}'] = to_tag
            ws[f'D{i}'] = comp.get('type', 'Cable')
            ws[f'E{i}'] = comp.get('cross_section', 0)
            ws[f'F{i}'] = conn.get('length', 0)
            ws[f'G{i}'] = comp.get('properties', {}).get('cores', 4)
            ws[f'H{i}'] = comp.get('voltage_rating', 0)
            ws[f'I{i}'] = conn.get('installation_method', 'E')
            ws[f'J{i}'] = comp.get('ampacity_base', 0)
            ws[f'K{i}'] = comp.get('conductor_material', 'Copper')
            ws[f'L{i}'] = comp.get('insulation_type', 'XLPE')
        
        self._auto_adjust_columns(ws)
    
    def export_calculation_worksheet(
        self,
        calc_type: str,
        results: Dict,
        project_data: Dict
    ) -> io.BytesIO:
        """
        Export calculation worksheet with formulas.
        
        Args:
            calc_type: Type of calculation (short_circuit, load_flow, etc.)
            results: Calculation results
            project_data: Project information
        
        Returns:
            BytesIO object containing Excel file
        """
        self.wb = Workbook()
        ws = self.wb.active
        ws.title = calc_type.replace('_', ' ').title()
        
        if calc_type == "short_circuit":
            self._create_short_circuit_worksheet(ws, results, project_data)
        elif calc_type == "load_flow":
            self._create_load_flow_worksheet(ws, results, project_data)
        elif calc_type == "arc_flash":
            self._create_arc_flash_worksheet(ws, results, project_data)
        
        output = io.BytesIO()
        self.wb.save(output)
        output.seek(0)
        
        return output
    
    def _create_short_circuit_worksheet(
        self,
        ws,
        results: Dict,
        project_data: Dict
    ):
        """Create short circuit calculation worksheet"""
        # Title
        ws['A1'] = "Short Circuit Calculations (IEC 60909)"
        ws['A1'].font = self.title_font
        
        ws['A2'] = f"Project: {project_data.get('name', 'N/A')}"
        ws['A3'] = f"Date: {datetime.now().strftime('%Y-%m-%d')}"
        
        # Input parameters
        row = 5
        ws[f'A{row}'] = "Input Parameters"
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Base MVA:"
        ws[f'B{row}'] = project_data.get('base_mva', 100)
        
        row += 1
        ws[f'A{row}'] = "System Frequency (Hz):"
        ws[f'B{row}'] = project_data.get('system_frequency', 50)
        
        # Results table
        row += 2
        headers = ['Bus Tag', 'I"k3 (kA)', 'ip (kA)', 'Ib (kA)', 'Sk (MVA)', 'Breaker Status']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
        
        # Data
        for bus_id, data in results.items():
            row += 1
            ws.cell(row=row, column=1, value=data.get('tag', bus_id))
            ws.cell(row=row, column=2, value=data.get('i_k3', 0))
            ws.cell(row=row, column=3, value=data.get('ip', 0))
            ws.cell(row=row, column=4, value=data.get('ib', 0))
            ws.cell(row=row, column=5, value=data.get('sk_mva', 0))
            ws.cell(row=row, column=6, value=data.get('breaker_status', 'N/A'))
        
        self._auto_adjust_columns(ws)
    
    def _create_arc_flash_worksheet(
        self,
        ws,
        results: Dict,
        project_data: Dict
    ):
        """Create arc flash calculation worksheet"""
        ws['A1'] = "Arc Flash Analysis (IEEE 1584-2018)"
        ws['A1'].font = self.title_font
        
        ws['A2'] = f"Project: {project_data.get('name', 'N/A')}"
        
        # Results table
        row = 4
        headers = ['Bus Tag', 'IE (cal/cmÂ²)', 'AFB (ft)', 'PPE Cat', 'Hazard Level']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
        
        for bus_id, data in results.items():
            row += 1
            ws.cell(row=row, column=1, value=data.get('tag', bus_id))
            ws.cell(row=row, column=2, value=data.get('incident_energy', 0))
            ws.cell(row=row, column=3, value=data.get('arc_flash_boundary', 0))
            ws.cell(row=row, column=4, value=data.get('ppe_category', 0))
            ws.cell(row=row, column=5, value=data.get('hazard_level', 'N/A'))
        
        self._auto_adjust_columns(ws)
    
    def _write_headers(self, ws, headers: List[str]):
        """Write styled headers to worksheet"""
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
    
    def _apply_header_style(self, ws, row: int, cols: int):
        """Apply header style to a row"""
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths based on content"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width


# Testing and example usage
if __name__ == "__main__":
    print("ðŸ“Š Excel Export Test")
    print("=" * 70)
    
    if not OPENPYXL_AVAILABLE:
        print("âŒ openpyxl not available - test skipped")
        print("Install with: pip install openpyxl")
    else:
        exporter = ExcelExporter()
        
        # Test project data
        test_project = {
            'name': 'Test Project',
            'base_mva': 100.0,
            'system_frequency': 50,
            'nodes': [
                {'id': 1, 'type': 'Source', 'custom_tag': 'SOURCE-01', 'component_library_id': 1, 'properties': {}},
                {'id': 2, 'type': 'Breaker', 'custom_tag': 'CB-01', 'component_library_id': 2, 'properties': {'poles': 3}},
                {'id': 3, 'type': 'Transformer', 'custom_tag': 'XFMR-01', 'component_library_id': 3, 'properties': {'rating_kva': 1000}},
            ],
            'connections': [
                {'source_node_id': 1, 'target_node_id': 2, 'cable_library_id': 10, 'length': 50},
                {'source_node_id': 2, 'target_node_id': 3, 'cable_library_id': 10, 'length': 100}
            ]
        }
        
        test_components = {
            '1': {'manufacturer': 'Utility', 'model': 'Grid', 'voltage_rating': 11},
            '2': {'manufacturer': 'Schneider', 'model': 'EasyPact', 'voltage_rating': 11, 'ampacity_base': 1000, 'short_circuit_rating': 50},
            '3': {'manufacturer': 'ABB', 'model': 'DryFlex', 'voltage_rating': 11, 'impedance_z_percent': 6},
            '10': {'type': 'Cable', 'manufacturer': 'Nexans', 'cross_section': 185, 'voltage_rating': 11, 'ampacity_base': 420, 'conductor_material': 'Copper', 'insulation_type': 'XLPE'}
        }
        
        print("\nâœ… Test data created")
        
        # Test equipment list export
        print("\nðŸ“„ Exporting equipment list...")
        try:
            excel_data = exporter.export_equipment_list(test_project, test_components)
            print(f"âœ… Generated Excel file: {len(excel_data.getvalue())} bytes")
            
            # Optionally save to file
            with open('/tmp/test_equipment_list.xlsx', 'wb') as f:
                f.write(excel_data.getvalue())
            print("âœ… Saved to /tmp/test_equipment_list.xlsx")
        except Exception as e:
            print(f"âŒ Export failed: {e}")
        
        print("\nâœ… Excel Export test complete!")
